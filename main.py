from flask import Flask, request, render_template, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
import os

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configuration
EXCEL_FILE = 'WMU Stuedents Upgrade 1.xlsx'
SHEET_NAME = 'MICHIGAN STUDENTS DATA'
HEADER_ROW = 3  # The actual headers are in row 3

def find_student(nama):
    """
    Find student by name (case-insensitive)
    Returns tuple: (row_number, exists)
    Column C (index 2) contains 'IDN'
    Column D (index 3) contains 'Nama'
    """
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    nama_lower = nama.lower().strip()

    # Start from row 4 (first data row after header in row 3)
    for idx in range(4, ws.max_row + 1):
        cell_value = ws.cell(row=idx, column=4).value  # Column D (Nama)
        if cell_value and cell_value.lower().strip() == nama_lower:
            wb.close()
            return idx, True

    wb.close()
    return None, False

def get_next_idn():
    """Get the next IDN number"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    max_idn = 0
    # Start from row 4 (first data row)
    for idx in range(4, ws.max_row + 1):
        idn_value = ws.cell(row=idx, column=3).value  # Column C (IDN)
        if idn_value and isinstance(idn_value, (int, float)):
            max_idn = max(max_idn, int(idn_value))

    wb.close()
    return max_idn + 1

def update_or_add_student(data):
    """
    Update existing student or add new one
    Returns: dict with status and message
    Column mapping:
    C (3) = IDN
    D (4) = Nama
    E (5) = Jurusan
    F (6) = University
    G (7) = Year
    H (8) = Provinsi
    """
    nama = data.get('nama', '').strip()
    jurusan = data.get('jurusan', '').strip()
    university = data.get('university', '').strip()
    year = data.get('year', '').strip()
    provinsi = data.get('provinsi', '').strip()

    if not nama:
        return {'status': 'error', 'message': 'Nama is required'}

    row_num, exists = find_student(nama)
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    if exists:
        # Update existing row (columns C through H)
        ws.cell(row=row_num, column=4, value=nama)  # Nama
        ws.cell(row=row_num, column=5, value=jurusan)  # Jurusan
        ws.cell(row=row_num, column=6, value=university)  # University
        ws.cell(row=row_num, column=7, value=year)  # Year
        ws.cell(row=row_num, column=8, value=provinsi)  # Provinsi

        wb.save(EXCEL_FILE)
        wb.close()
        return {
            'status': 'updated',
            'message': f'Successfully updated record for {nama}',
            'data': {
                'nama': nama,
                'jurusan': jurusan,
                'university': university,
                'year': year,
                'provinsi': provinsi
            }
        }
    else:
        # Add new row at the end
        new_row = ws.max_row + 1
        idn = get_next_idn()

        ws.cell(row=new_row, column=3, value=idn)  # IDN
        ws.cell(row=new_row, column=4, value=nama)  # Nama
        ws.cell(row=new_row, column=5, value=jurusan)  # Jurusan
        ws.cell(row=new_row, column=6, value=university)  # University
        ws.cell(row=new_row, column=7, value=year)  # Year
        ws.cell(row=new_row, column=8, value=provinsi)  # Provinsi

        wb.save(EXCEL_FILE)
        wb.close()
        return {
            'status': 'added',
            'message': f'Successfully added new record for {nama}',
            'data': {
                'idn': idn,
                'nama': nama,
                'jurusan': jurusan,
                'university': university,
                'year': year,
                'provinsi': provinsi
            }
        }

@app.route('/')
def index():
    """Render the input form"""
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    """Handle form submission"""
    try:
        data = request.form.to_dict()
        print(f"Received data: {data}")  # Debug log
        result = update_or_add_student(data)
        print(f"Result: {result}")  # Debug log
        return jsonify(result)
    except Exception as e:
        import traceback
        error_msg = str(e)
        print(f"Error occurred: {error_msg}")  # Debug log
        traceback.print_exc()  # Print full traceback
        return jsonify({'status': 'error', 'message': error_msg}), 500

@app.route('/api/submit', methods=['POST'])
def api_submit():
    """API endpoint for JSON submissions"""
    try:
        data = request.get_json()
        result = update_or_add_student(data)
        return jsonify(result)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
