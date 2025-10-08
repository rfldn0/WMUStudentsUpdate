from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('WMU Stuedents Upgrade 1.xlsx')

# Print sheet names
print("Sheet names:", wb.sheetnames)
print()

# Check if 'MICHIGAN STUDENTS DATA' exists
if 'MICHIGAN STUDENTS DATA' in wb.sheetnames:
    ws = wb['MICHIGAN STUDENTS DATA']

    print("Headers (Row 1):")
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
        print(f"  - {cell.value}")

    print(f"\nTotal rows with data: {ws.max_row}")
    print(f"Total columns: {ws.max_column}")

    print("\nFirst 3 data rows:")
    for row_idx in range(2, min(5, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row=row_idx, column=col_idx).value)
        print(f"  Row {row_idx}: {row_data}")
else:
    print("'MICHIGAN STUDENTS DATA' sheet not found!")
    print("Available sheets:", wb.sheetnames)
